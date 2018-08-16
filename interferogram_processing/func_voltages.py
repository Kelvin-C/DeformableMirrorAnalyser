# |**********************************************************************;
# * Project           : MSci Project: PLAS-Smith-3
# *
# * Program name      : func_voltages.py
# *
# * Author            : Kelvin Chan
# *
# * Date created      : 09 Dec 2017
# *
# * Purpose           : Reads the voltages off the .xlsx spreadsheet.
# *
# * Revision History  : v1.0
# *
# |**********************************************************************;

from openpyxl import load_workbook

def get_row(sheet, row):
    """
    Output the elements of a specific row from a specific sheet in the .xlsx file.
    """
    elements = []
    output = 1
    i = 0
    while output is not None:
        i += 1
        output = sheet.cell(row=row, column=i).value
        elements.append(output)

    elements = elements[:-1]
    return elements

def get_column(sheet, column, include_heading = False):
    """
    Output the elements of a specific column from a specific sheet in the .xlsx file.
    """
    elements = []
    output = 1
    i = 0
    while output is not None:
        i += 1
        output = sheet.cell(row=i, column=column).value
        elements.append(output)

    if include_heading is False:
        elements = elements[1:-1]
    else:
        elements = elements[:-1]
    return elements

def get_table(sheet, include_heading = False):
    """
    Output the all elements from a specific sheet in the .xlsx file.
    """
    headings = get_row(sheet, row=1)
    elements = []
    for i in range(len(headings)):
        elements.append([])
        elements[i] = get_column(sheet, column=i+1, include_heading=include_heading)
    return elements


def get_voltages(xlsx_file):
    """
    Algorithm to output the list of voltages from the .xslsx file.
    """
    wb = load_workbook(xlsx_file)
    sheetname_list = wb.get_sheet_names()
    sheet_list = list(map(lambda i: wb.get_sheet_by_name(sheetname_list[i]),
                          range(len(sheetname_list))))  # make a list of sheet objects

    sheet_elements_list = []

    for i in range(len(sheet_list)):
        sheet_elements_list.append(get_table(sheet_list[i]))

    return sheet_elements_list[0][0]