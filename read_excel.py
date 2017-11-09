from openpyxl import load_workbook
import matplotlib.pyplot as plt

#Locate '.xslx' file here
xlsx_dir = 'xslx_files/'
xlsx_filename = 'py_circuits.xlsx'
xlsx_file = xlsx_dir + xlsx_filename

wb = load_workbook(xlsx_file)
sheetname_list = wb.get_sheet_names()
sheet_list = map(lambda i: wb.get_sheet_by_name(sheetname_list[i]), range(len(sheetname_list)))   #make a list of sheet objects

def get_row(sheet, row):
    elements = []
    output = 1
    i = 0
    while output is not None:
        i += 1
        output = sheet.cell(row=row, column=i).value
        elements.append(output)

    elements = elements[:-1]
    return elements

def get_column(sheet, column, include_heading = False):
    elements = []
    output = 1
    i = 0
    while output is not None:
        i += 1
        output = sheet.cell(row=i, column=column).value
        elements.append(output)

    if include_heading is False:
        elements = elements[1:-1]
    else:
        elements = elements[:-1]
    return elements

def get_table(sheet, include_heading = False):
    headings = get_row(sheet, row=1)
    elements = []
    for i in range(len(headings)):
        elements.append([])
        elements[i] = get_column(sheet, column=i+1, include_heading=include_heading)
    return elements

sheet_elements_list = []

for i in range(len(sheet_list)):
    sheet_elements_list.append(get_table(sheet_list[i]))

fig = plt.figure(1)

sheet_plot_indices = [0, 1, 2]

plt.subplot(1,2,1)
for i in sheet_plot_indices:
    plt.plot(sheet_elements_list[i][0], sheet_elements_list[i][1], label=sheetname_list[i])
plt.ylabel('Average of Voltage Oscillation at Steady State $(V)$')
plt.xlabel('Number of LRC/RC Circuits')
plt.legend()

plt.subplot(1,2,2)
for i in sheet_plot_indices:
    plt.plot(sheet_elements_list[i][0], sheet_elements_list[i][2], label=sheetname_list[i])
plt.ylabel('Peak-to-Peak Voltage Difference at Steady State $(V)$')
plt.xlabel('Number of LRC/RC Circuits')
plt.legend()

plt.show()




